from openpyxl import Workbook, load_workbook

rd=load_workbook(filename=r"C:/Users/suri1/OneDrive/Documents/python_automation/automation_practise/FileIODemo/demoexcelsheet.xlsx")
ws=rd["Sheet"]

print(ws["D5"].value)
